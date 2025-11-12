import streamlit as st
import pandas as pd
import io
from collections import OrderedDict
import re
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Convertir TXT Perceptron a Excel", layout="wide")

# --- ESTILO GLOBAL (FONDO OSCURO, TABLAS CLARAS) ---
st.markdown(""" 
    <style>
    body {
        background-color: #121212;
        color: #FFFFFF;
        font-family: 'Poppins', sans-serif;
    }

    .stApp {
        background-color: #121212;
    }

    /* Encabezados */
    h1, h2, h3, h4 {
        color: #ffc107;
    }

    /* Área de subida de archivos */
    div[data-testid="stFileUploader"] {
        border: 2px dashed #5a5a5a !important;
        background-color: rgba(50,50,50,0.7);
        border-radius: 15px;
        padding: 20px;
    }

    div[data-testid="stFileUploader"]:hover {
        border-color: #ffc107 !important;
        background-color: rgba(80,80,80,0.9);
    }

    /* Tabla de correlación */
    .dataframe {
        background: #2b2b2b !important; /* gris oscuro pero más claro que el fondo */
        color: #ffffff !important;
        border-radius: 10px;
        font-size: 15px;
    }

    .dataframe td, .dataframe th {
        text-align: center !important;
        padding: 8px !important;
    }

    /* Botón de descarga */
    div.stDownloadButton > button {
        background-color: #ffc107;
        color: #000;
        font-weight: bold;
        border-radius: 10px;
        border: none;
        padding: 10px 25px;
    }

    div.stDownloadButton > button:hover {
        background-color: #ffde59;
        color: #000;
    }
    </style>
""", unsafe_allow_html=True)

# --- TÍTULO ---
st.title("📄 Comparativo Frontal vs Final")

# --- FUNCIÓN PARA PROCESAR TXT ---
def procesar_txt_a_df(archivo):
    contenido = archivo.read().decode("latin-1").splitlines()
    encabezados, mediciones = [], []

    for linea in contenido:
        partes = linea.strip().split("\t")
        if "JSN" in partes and "PSN" in partes:
            encabezados = partes
            continue
        if partes and partes[0].upper() not in ["NOMINAL", "USL", "LSL", "UTL", "LTL", "URL", "LRL"]:
            mediciones.append(partes)

    if not encabezados or not mediciones:
        return None, []

    filas_med = []
    for med in mediciones:
        fila = OrderedDict({
            "JSN": med[0],
            "PSN": med[1] if len(med) > 1 else "",
            "Fecha": med[2] if len(med) > 2 else "",
            "Hora": med[3] if len(med) > 3 else ""
        })
        for i, col in enumerate(encabezados[4:], start=4):
            fila[col] = med[i] if i < len(med) else ""
        filas_med.append(fila)

    eje_cols = encabezados[4:]
    return pd.DataFrame(filas_med), eje_cols

# --- MAPEO DE EJES ---
def map_axis(front_axis):
    match = re.match(r"(1100)([LR]\[[XYZ]\])", front_axis)
    if match:
        return f"3125{match.group(2)}"
    return front_axis

# --- SUBIDA DE ARCHIVOS ---
st.subheader("📤 Archivo Frontal/Trasero")
archivo_frontal = st.file_uploader("Carga el archivo TXT Frontal de Perceptron", type=["txt"], key="frontal")

st.subheader("📤 Archivo Final")
archivo_final = st.file_uploader("Carga el archivo TXT Final de Perceptron", type=["txt"], key="final")

# --- PROCESAMIENTO ---
if archivo_frontal and archivo_final:
    df_frontal, ejes_frontal = procesar_txt_a_df(archivo_frontal)
    df_final, ejes_final = procesar_txt_a_df(archivo_final)

    if df_frontal is None or df_final is None:
        st.error("⚠️ Uno de los archivos no contiene mediciones reales válidas.")
    else:
        st.success("✅ Archivos procesados correctamente. Descarga combinada lista.")

        # --- MATCH DE PSN (solo los que coinciden exactamente) ---
        df_match = pd.DataFrame({
            "FrontJSN": df_frontal["PSN"],
            "FinalJSN": df_final["PSN"]
        })

        # 🔹 Filtrar solo PSN iguales (coinciden frontal y final)
        df_match = df_match[df_match["FrontJSN"] == df_match["FinalJSN"]].copy()

        # 🔹 Tomar esos PSN válidos
        psn_validos = df_match["FrontJSN"].unique()

        # 🔹 Filtrar los DataFrames originales según PSN coincidentes
        df_frontal = df_frontal[df_frontal["PSN"].isin(psn_validos)]
        df_final = df_final[df_final["PSN"].isin(psn_validos)]

        ejes_mapeados = []
        for eje in ejes_frontal:
            eje_final = map_axis(eje)
            if eje_final in ejes_final or eje_final.startswith("3125"):
                ejes_mapeados.append((eje, eje_final))
        df_axes = pd.DataFrame(ejes_mapeados, columns=["Front-Axis", "Final-Axis"])

        # --- Cálculo de correlaciones con emparejamiento por PSN ---
        correlacion_data = []

        # Unimos frontal y final por PSN (solo los que coincidan)
        df_merge = pd.merge(
            df_frontal, df_final,
            on="PSN", suffixes=("_front", "_final")
        )

        for front_eje, final_eje in df_axes.values:
            col_front = f"{front_eje}_front" if f"{front_eje}_front" in df_merge.columns else front_eje
            col_final = f"{final_eje}_final" if f"{final_eje}_final" in df_merge.columns else final_eje

            if col_front in df_merge.columns and col_final in df_merge.columns:
                front_vals = pd.to_numeric(df_merge[col_front], errors="coerce")
                final_vals = pd.to_numeric(df_merge[col_final], errors="coerce")

                # 🔹 Eliminar NaN y alinear índices
                valid_idx = front_vals.dropna().index.intersection(final_vals.dropna().index)
                front_vals = front_vals.loc[valid_idx].reset_index(drop=True)
                final_vals = final_vals.loc[valid_idx].reset_index(drop=True)

                # 🔹 Evitar error si los tamaños no coinciden
                min_len = min(len(front_vals), len(final_vals))
                if min_len < 2:
                    continue  # saltar si no hay suficientes valores

                front_vals = front_vals.iloc[:min_len]
                final_vals = final_vals.iloc[:min_len]

                # --- Cálculos ---
                front_mean = np.mean(front_vals)
                final_mean = np.mean(final_vals)
                correlation = np.corrcoef(front_vals, final_vals)[0, 1]
                sigma6 = np.std(front_vals) * 6
                offset_calc = front_mean - final_mean

                correlacion_data.append([
                    front_eje, final_eje,
                    round(front_mean, 3),
                    round(final_mean, 3),
                    round(correlation, 3),
                    round(sigma6, 3),
                    round(offset_calc, 3)
                ])



        # --- CONVERTIR LOS DATOS DE CORRELACIÓN A DATAFRAME ---
        df_correlacion = pd.DataFrame(
            correlacion_data,
            columns=[
                "Front-Axis", "Final-Axis",
                "Front-Mean", "Final-Mean",
                "Correlation", "6Sigma", "Calculated-Offset"
            ]
        )
        # Agregar columna para selección
        df_correlacion["Punto"] = df_correlacion["Front-Axis"]

        # --- FUNCIONES DE COLORES ---
        def colorear_correlacion(val):
            if isinstance(val, (int, float)):
                if val >= 0.7:
                    return 'background-color: #47FF47; color: #000000; font-weight: 600;'  # Verde
                elif val >= 0.69:
                    return 'background-color: #FFFD00; color: #000000; font-weight: 600;'  # Amarillo
            return 'color: #FFFFFF;'  # Texto blanco normal

        def colorear_offset(val):
            if isinstance(val, (int, float)):
                if abs(val) > 1:
                    return 'background-color: #FF0000; color: #FFFFFF; font-weight: 600;'  # Rojo
                elif abs(val) > 0.5:
                    return 'background-color: #FFFD00; color: #000000; font-weight: 600;'  # Amarillo
            return 'color: #FFFFFF;'  # Texto blanco por defecto

        # --- APLICAR ESTILO AL DATAFRAME ---
        df_correlacion_styled = (
            df_correlacion.style
            .applymap(colorear_correlacion, subset=["Correlation"])
            .applymap(colorear_offset, subset=["Calculated-Offset"])
            .set_table_styles([
                {'selector': 'th', 'props': [('background-color', '#2b2b2b'),
                                            ('color', '#FFFFFF'),
                                            ('font-weight', 'bold'),
                                            ('text-align', 'center'),
                                            ('padding', '8px')]},
                {'selector': 'td', 'props': [('background-color', '#1e1e1e'),
                                            ('color', '#FFFFFF'),
                                            ('text-align', 'center'),
                                            ('padding', '8px')]},
                {'selector': 'tbody tr:hover', 'props': [('background-color', '#333333')]},
                {'selector': 'table', 'props': [('border-radius', '10px'),
                                                ('overflow', 'hidden'),
                                                ('border', '1px solid #444')]}
            ])
        )

        # --- MOSTRAR EN STREAMLIT ---
        st.subheader("📈 Correlación")
        st.dataframe(df_correlacion_styled, use_container_width=True)

        # --- SELECCIÓN DE PUNTOS PARA XML ---
        st.subheader("✨ Selecciona los puntos que quieres incluir en el XML")

        puntos_disponibles = df_correlacion["Punto"].unique()

        puntos_seleccionados = st.multiselect(
            "Puntos disponibles:",
            options=puntos_disponibles,
            default=puntos_disponibles  # Por defecto selecciona todos
        )

        # Filtrar según selección
        df_filtrado = df_correlacion[df_correlacion["Punto"].isin(puntos_seleccionados)]


        # --- BOTÓN DE DESCARGA CON ESTILOS EN EXCEL ---
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_frontal.to_excel(writer, index=False, sheet_name="Frontal")
            df_final.to_excel(writer, index=False, sheet_name="Final")
            df_match.to_excel(writer, index=False, sheet_name="Match_PSN")
            df_axes.to_excel(writer, index=False, sheet_name="Eje-Mapping")
            df_correlacion.to_excel(writer, index=False, sheet_name="Correlacion")
            #writer.save()

        # Abrir el libro para aplicar colores en la hoja "Correlacion"
        buffer.seek(0)  # volver al inicio del buffer
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Correlacion"]

        # Colores para Correlation
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):  # Columna "Correlation"
            for cell in row:
                if cell.value is not None:
                    if cell.value >= 0.7:
                        cell.fill = PatternFill(start_color="47FF47", end_color="47FF47", fill_type="solid")
                    elif cell.value >= 0.69:
                        cell.fill = PatternFill(start_color="FFFD00", end_color="FFFD00", fill_type="solid")

        # Colores para Calculated-Offset
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):  # Columna "Calculated-Offset"
            for cell in row:
                if cell.value is not None:
                    if abs(cell.value) > 1:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    elif abs(cell.value) > 0.5:
                        cell.fill = PatternFill(start_color="FFFD00", end_color="FFFD00", fill_type="solid")

        # Sombreado ligero para Front-Mean (col 3) y Final-Mean (col 4)
        mean_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # gris claro
        for col in [3, 4]:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    if cell.value is not None:
                        cell.fill = mean_fill

        # Guardar en un nuevo buffer para descargar
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        st.download_button(
            label="📥 Descargar Excel completo coloreado",
            data=excel_buffer,
            file_name="Mediciones_Percepton_Completo_Coloreado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# --- BOTÓN DE DESCARGA EN FORMATO XML (agrupando por checkpoint) ---
# --- BOTÓN DE DESCARGA EN FORMATO XML (agrupando por checkpoint) ---
        def generar_xml_comparacion(df, station_name="T1XX_SUV_Front_Mod", model_name="K_SUV"):
            import xml.etree.ElementTree as ET
            
            # Nodo raíz
            gauge = ET.Element("GAUGE")
            station = ET.SubElement(gauge, "STATION")
            ET.SubElement(station, "NAME").text = station_name
            model = ET.SubElement(station, "MODEL")
            ET.SubElement(model, "NAME").text = model_name
            
            # Agrupar por checkpoint (extraemos nombre antes de [X], [Y], [Z])
            df["Checkpoint"] = df["Front-Axis"].str.extract(r"(^\d+[LR])")  # Ej: 1100L
            df["Axis"] = df["Front-Axis"].str.extract(r"\[([XYZ])\]")        # Ej: X, Y, Z
            
            for checkpoint_name, group in df.groupby("Checkpoint"):
                checkpoint = ET.SubElement(model, "CHECKPOINT")
                ET.SubElement(checkpoint, "NAME").text = checkpoint_name
                
                # Agregar AXIS X, Y, Z según Calculated-Offset
                for axis in ["X", "Y", "Z"]:
                    axis_node = ET.SubElement(checkpoint, "AXIS")
                    ET.SubElement(axis_node, "NAME").text = axis
                    val = group.loc[group["Axis"]==axis, "Calculated-Offset"]
                    ET.SubElement(axis_node, "OFFSET").text = str(round(val.values[0],3)) if not val.empty else "0"
                
                # Agregar Diameter siempre 0
                axis_node = ET.SubElement(checkpoint, "AXIS")
                ET.SubElement(axis_node, "NAME").text = "Diameter"
                ET.SubElement(axis_node, "OFFSET").text = "0"
            
            # Convertir a string
            xml_str = ET.tostring(gauge, encoding="utf-8", method="xml")
            return xml_str

        # Generar XML
        xml_data = generar_xml_comparacion(df_filtrado)


        # Botón de descarga
        st.download_button(
            label="📥 Descargar comparación en XML",
            data=xml_data,
            file_name="Comparacion_Percepton.xml",
            mime="application/xml"

        )
